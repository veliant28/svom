from __future__ import annotations

from datetime import datetime, timezone as dt_timezone
from pathlib import Path
from typing import Iterable

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.utils.translation import gettext as _
from openpyxl import Workbook

from apps.autocatalog.models import CarMake, CarModel, CarModification, UtrArticleDetailMap, UtrDetailCarMap

MAX_EXCEL_DATA_ROWS_PER_SHEET = 1_000_000


class Command(BaseCommand):
    help = _(
        "Экспортирует полный бэкап автокаталога в XLSX "
        "(марки, модели, модификации, UTR mapping-и, связи и restore-план)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            type=str,
            default="",
            help=_("Путь к XLSX файлу. По умолчанию: <repo_root>/autocatalog_backup_<timestamp>.xlsx"),
        )
        parser.add_argument(
            "--with-detail-matrix",
            action="store_true",
            help=_("Добавить тяжелый аналитический лист detail_dependency_matrix (может сильно замедлить экспорт)."),
        )

    def handle(self, *args, **options):
        output_path = self._resolve_output_path(raw_output=str(options.get("output") or "").strip())
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook(write_only=True)
        row_counts: dict[str, int] = {}

        self.stdout.write("Preparing sheet: restore_plan ...")
        row_counts["restore_plan"] = self._append_restore_plan_sheet(
            workbook,
            include_detail_matrix=bool(options.get("with_detail_matrix")),
        )
        self.stdout.write(f"Done: restore_plan ({row_counts['restore_plan']})")
        self.stdout.write("Preparing sheet: car_makes ...")
        row_counts["car_makes"] = self._append_sheet(
            workbook=workbook,
            title="car_makes",
            header=(
                "id",
                "name",
                "slug",
                "normalized_name",
                "created_at",
                "updated_at",
            ),
            rows=(
                (
                    make.id,
                    make.name,
                    make.slug,
                    make.normalized_name,
                    self._fmt_dt(make.created_at),
                    self._fmt_dt(make.updated_at),
                )
                for make in CarMake.objects.order_by("id").iterator(chunk_size=2000)
            ),
        )
        self.stdout.write(f"Done: car_makes ({row_counts['car_makes']})")
        self.stdout.write("Preparing sheet: car_models ...")
        row_counts["car_models"] = self._append_sheet(
            workbook=workbook,
            title="car_models",
            header=(
                "id",
                "make_id",
                "make_name",
                "make_slug",
                "name",
                "slug",
                "normalized_name",
                "created_at",
                "updated_at",
            ),
            rows=(
                (
                    model.id,
                    model.make_id,
                    model.make.name,
                    model.make.slug,
                    model.name,
                    model.slug,
                    model.normalized_name,
                    self._fmt_dt(model.created_at),
                    self._fmt_dt(model.updated_at),
                )
                for model in CarModel.objects.select_related("make").order_by("id").iterator(chunk_size=2000)
            ),
        )
        self.stdout.write(f"Done: car_models ({row_counts['car_models']})")
        self.stdout.write("Preparing sheet: car_modifications ...")
        row_counts["car_modifications"] = self._append_sheet(
            workbook=workbook,
            title="car_modifications",
            header=(
                "id",
                "make_id",
                "make_name",
                "make_slug",
                "model_id",
                "model_name",
                "model_slug",
                "start_date_at",
                "end_date_at",
                "year",
                "modification",
                "capacity",
                "engine",
                "hp_from",
                "kw_from",
                "dedupe_key",
                "created_at",
                "updated_at",
            ),
            rows=(
                (
                    modification.id,
                    modification.make_id,
                    modification.make.name,
                    modification.make.slug,
                    modification.model_id,
                    modification.model.name,
                    modification.model.slug,
                    self._fmt_date(modification.start_date_at),
                    self._fmt_date(modification.end_date_at),
                    modification.year,
                    modification.modification,
                    modification.capacity,
                    modification.engine,
                    modification.hp_from,
                    modification.kw_from,
                    modification.dedupe_key,
                    self._fmt_dt(modification.created_at),
                    self._fmt_dt(modification.updated_at),
                )
                for modification in CarModification.objects.select_related("make", "model")
                .order_by("id")
                .iterator(chunk_size=2000)
            ),
        )
        self.stdout.write(f"Done: car_modifications ({row_counts['car_modifications']})")
        self.stdout.write("Preparing sheet: utr_article_detail_maps ...")
        row_counts["utr_article_detail_maps"] = self._append_sheet(
            workbook=workbook,
            title="utr_article_detail_maps",
            header=(
                "id",
                "article",
                "normalized_article",
                "brand_name",
                "normalized_brand",
                "utr_detail_id",
                "created_at",
                "updated_at",
            ),
            rows=(
                (
                    item.id,
                    item.article,
                    item.normalized_article,
                    item.brand_name,
                    item.normalized_brand,
                    item.utr_detail_id,
                    self._fmt_dt(item.created_at),
                    self._fmt_dt(item.updated_at),
                )
                for item in UtrArticleDetailMap.objects.order_by("id").iterator(chunk_size=5000)
            ),
        )
        self.stdout.write(f"Done: utr_article_detail_maps ({row_counts['utr_article_detail_maps']})")
        self.stdout.write("Preparing sheet(s): utr_detail_car_maps_* ...")
        total_rows, chunk_count = self._append_utr_detail_car_map_chunks(workbook=workbook, row_counts=row_counts)
        self.stdout.write(f"Done: utr_detail_car_maps_* ({total_rows} rows across {chunk_count} sheet(s))")
        if options.get("with_detail_matrix"):
            self.stdout.write("Preparing sheet: detail_dependency_matrix ...")
            row_counts["detail_dependency_matrix"] = self._append_sheet(
                workbook=workbook,
                title="detail_dependency_matrix",
                header=(
                    "utr_detail_id",
                    "article_count",
                    "vehicle_count",
                    "sample_articles",
                    "sample_vehicles",
                ),
                rows=self._iter_detail_dependency_rows(),
            )
            self.stdout.write(f"Done: detail_dependency_matrix ({row_counts['detail_dependency_matrix']})")

        self.stdout.write("Preparing sheet: manifest ...")
        row_counts["manifest"] = self._append_manifest_sheet(workbook=workbook, output_path=output_path, row_counts=row_counts)
        self.stdout.write(f"Done: manifest ({row_counts['manifest']})")
        workbook.save(output_path)

        self.stdout.write(self.style.SUCCESS(f"Autocatalog XLSX backup created: {output_path}"))
        for sheet_name, count in row_counts.items():
            self.stdout.write(f"  - {sheet_name}: {count}")

    def _resolve_output_path(self, *, raw_output: str) -> Path:
        if raw_output:
            output_path = Path(raw_output)
            if not output_path.is_absolute():
                output_path = (Path(settings.BASE_DIR) / output_path).resolve()
            return output_path

        timestamp = datetime.now(tz=dt_timezone.utc).strftime("%Y%m%d_%H%M%S")
        repo_root = Path(settings.BASE_DIR).parent
        return (repo_root / f"autocatalog_backup_{timestamp}.xlsx").resolve()

    def _append_restore_plan_sheet(self, workbook: Workbook, *, include_detail_matrix: bool) -> int:
        rows = [
            ("order", "sheet", "model", "pk_field", "fk_dependencies", "restore_notes"),
            (1, "car_makes", "autocatalog.CarMake", "id", "-", "Сначала марки."),
            (2, "car_models", "autocatalog.CarModel", "id", "make_id -> car_makes.id", "Затем модели."),
            (
                3,
                "car_modifications",
                "autocatalog.CarModification",
                "id",
                "make_id -> car_makes.id; model_id -> car_models.id",
                "Потом модификации.",
            ),
            (4, "utr_article_detail_maps", "autocatalog.UtrArticleDetailMap", "id", "-", "UTR article->detail mapping."),
            (
                5,
                "utr_detail_car_maps_<n>",
                "autocatalog.UtrDetailCarMap",
                "id",
                "car_modification_id -> car_modifications.id",
                "UTR detail->car relation в самом конце. Лист разбивается на chunks по лимиту Excel.",
            ),
        ]
        if include_detail_matrix:
            rows.append(
                (
                    6,
                    "detail_dependency_matrix",
                    "derived",
                    "-",
                    "-",
                    "Диагностический лист: быстрое понимание связности detail_id.",
                )
            )
        sheet = workbook.create_sheet("restore_plan")
        count = 0
        for row in rows:
            sheet.append(row)
            count += 1
        return count - 1

    def _append_manifest_sheet(self, *, workbook: Workbook, output_path: Path, row_counts: dict[str, int]) -> int:
        sheet = workbook.create_sheet("manifest")
        rows = (
            ("generated_at_utc", datetime.now(tz=dt_timezone.utc).isoformat()),
            ("project_base_dir", str(Path(settings.BASE_DIR))),
            ("file_path", str(output_path)),
            ("note", "This workbook is designed as XLSX backup + restore helper for autocatalog domain."),
            ("table", "rows"),
            *tuple((name, count) for name, count in sorted(row_counts.items())),
        )
        count = 0
        for row in rows:
            sheet.append(row)
            count += 1
        return count - 5

    def _append_sheet(
        self,
        *,
        workbook: Workbook,
        title: str,
        header: tuple[str, ...],
        rows: Iterable[tuple],
    ) -> int:
        if len(title) > 31:
            raise CommandError(f"Worksheet title is too long for XLSX: {title}")
        sheet = workbook.create_sheet(title=title)
        sheet.append(header)
        count = 0
        for row in rows:
            sheet.append(row)
            count += 1
        return count

    def _append_utr_detail_car_map_chunks(self, *, workbook: Workbook, row_counts: dict[str, int]) -> tuple[int, int]:
        header = (
            "id",
            "utr_detail_id",
            "car_modification_id",
            "make_id",
            "make_name",
            "model_id",
            "model_name",
            "year",
            "modification",
            "capacity",
            "engine",
            "end_date_at",
            "created_at",
        )
        queryset = (
            UtrDetailCarMap.objects.select_related(
                "car_modification",
                "car_modification__make",
                "car_modification__model",
            )
            .order_by("id")
            .iterator(chunk_size=5000)
        )

        chunk_index = 1
        sheet_name = f"utr_detail_car_maps_{chunk_index}"
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(header)
        chunk_rows = 0
        total_rows = 0

        for item in queryset:
            if chunk_rows >= MAX_EXCEL_DATA_ROWS_PER_SHEET:
                row_counts[sheet_name] = chunk_rows
                chunk_index += 1
                sheet_name = f"utr_detail_car_maps_{chunk_index}"
                sheet = workbook.create_sheet(title=sheet_name)
                sheet.append(header)
                chunk_rows = 0

            row = (
                item.id,
                item.utr_detail_id,
                item.car_modification_id,
                item.car_modification.make_id,
                item.car_modification.make.name,
                item.car_modification.model_id,
                item.car_modification.model.name,
                item.car_modification.year,
                item.car_modification.modification,
                item.car_modification.capacity,
                item.car_modification.engine,
                self._fmt_date(item.car_modification.end_date_at),
                self._fmt_dt(item.created_at),
            )
            sheet.append(row)
            chunk_rows += 1
            total_rows += 1

        row_counts[sheet_name] = chunk_rows
        return total_rows, chunk_index

    def _iter_detail_dependency_rows(self) -> Iterable[tuple]:
        detail_ids = set(UtrArticleDetailMap.objects.values_list("utr_detail_id", flat=True))
        detail_ids.update(UtrDetailCarMap.objects.values_list("utr_detail_id", flat=True))

        article_lookup: dict[str, list[str]] = {}
        for article, brand, detail_id in UtrArticleDetailMap.objects.order_by("id").values_list(
            "article",
            "brand_name",
            "utr_detail_id",
        ):
            if not detail_id:
                continue
            article_lookup.setdefault(detail_id, []).append(f"{article} ({brand})" if brand else article)

        vehicle_lookup: dict[str, list[str]] = {}
        vehicle_rows = UtrDetailCarMap.objects.select_related(
            "car_modification",
            "car_modification__make",
            "car_modification__model",
        ).order_by("id")
        for item in vehicle_rows.iterator(chunk_size=5000):
            mod = item.car_modification
            vehicle_label = f"{mod.make.name} {mod.model.name} {mod.year or '-'} {mod.modification}".strip()
            vehicle_lookup.setdefault(item.utr_detail_id, []).append(vehicle_label)

        for detail_id in sorted(detail_ids):
            articles = article_lookup.get(detail_id, [])
            vehicles = vehicle_lookup.get(detail_id, [])
            yield (
                detail_id,
                len(articles),
                len(vehicles),
                "; ".join(articles[:5]),
                "; ".join(vehicles[:5]),
            )

    def _fmt_dt(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            if value.tzinfo is None:
                value = value.replace(tzinfo=dt_timezone.utc)
            return value.astimezone(dt_timezone.utc).isoformat()
        return str(value)

    def _fmt_date(self, value) -> str:
        if value is None:
            return ""
        return value.isoformat()
